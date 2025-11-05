from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, g
from datetime import date, datetime
from sqlalchemy import func
import calendar
from io import BytesIO
import openpyxl
from database import WorkDay, WorkSegment, db

from auth import login_required

work_bp = Blueprint('work', __name__, url_prefix='/work')

def _parse_segments_csv(csv_text: str):
    if not csv_text:
        return []
    items = [s.strip() for s in csv_text.split(',')]
    return [s for s in items if s]

def _upsert_segments(d: WorkDay, csv_text: str):
    names = _parse_segments_csv(csv_text)
    d.segments.clear()
    for i, name in enumerate(names):
        d.segments.append(WorkSegment(seq=i, location_name=name))

def _append_segments(d: WorkDay, csv_text: str):
    names = _parse_segments_csv(csv_text)
    start_seq = len(d.segments)
    for i, name in enumerate(names):
        d.segments.append(WorkSegment(seq=start_seq + i, location_name=name))

@work_bp.route('/list')
@login_required
def list():
    # Defaults to current month
    y = request.args.get('year', type=int, default=date.today().year)
    m = request.args.get('month', type=int, default=date.today().month)
    start_d = date(y, m, 1)
    last_day = calendar.monthrange(y, m)[1]
    end_d = date(y, m, last_day)

    days = (db.session.query(WorkDay)
            .filter(WorkDay.day >= start_d, WorkDay.day <= end_d)
            .order_by(WorkDay.day.desc())
            .all())

    rows = []
    for d in days:
        rows.append({
            'id': d.id,
            'day': d.day,
            'status': d.status,
            'start_odo': d.start_odo,
            'end_odo': d.end_odo,
            'total_miles': d.compute_total_miles(),
            'start_location': d.start_location or '',
            'segments_str': ' to '.join([s.location_name for s in d.segments]),
            'trip_explanation': d.trip_explanation or '',
            'created_at': d.created_at,
            'updated_at': d.updated_at,
        })

    return render_template('work/list.html', rows=rows, year=y, month=m)

@work_bp.route('/start', methods=['GET', 'POST'])
@login_required
def start():
    if request.method == 'POST':
        # Enforce one active Work day at a time
        active = WorkDay.query.filter_by(status='started').first()
        if active:
            flash('You already have a started Work Day. End it or edit it before starting a new one.', 'danger')
            return redirect(url_for('work.list'))

        day_str = request.form.get('day')
        day_val = datetime.strptime(day_str, '%Y-%m-%d').date() if day_str else date.today()

        start_odo_raw = request.form.get('start_odo')
        start_odo_val = int(start_odo_raw) if start_odo_raw else None  # optional for backfill

        d = WorkDay(
            day=day_val,
            status='started',
            start_odo=start_odo_val,
            start_location=request.form.get('start_location') or None,
            trip_explanation=request.form.get('trip_explanation') or None,
        )

        segments_csv = request.form.get('segments_csv', '')
        _upsert_segments(d, segments_csv)

        db.session.add(d)
        db.session.commit()
        flash('Work Day started.', 'success')
        return redirect(url_for('work.list'))

    return render_template('work/start.html', today=date.today())

@work_bp.route('/update/<int:day_id>', methods=['GET', 'POST'])
@login_required
def update(day_id):
    d = WorkDay.query.get_or_404(day_id)
    if request.method == 'POST':
        _append_segments(d, request.form.get('append_segments', ''))
        d.trip_explanation = request.form.get('trip_explanation') or d.trip_explanation
        d.start_location = request.form.get('start_location')
        if request.form.get('start_odo'):
            d.start_odo = int(request.form['start_odo'])
        db.session.commit()
        flash('Work Day updated successfully.', 'success')
        return redirect(url_for('work.list'))
    return render_template('work/update.html', d=d)

@work_bp.route('/view/<int:day_id>', methods=['GET', 'POST'])
@login_required
def view(day_id):
    d = WorkDay.query.get_or_404(day_id)
    if request.method == 'POST':
        # Full edit/backfill — all fields editable
        day_in = request.form.get('day')
        d.day = datetime.strptime(day_in, '%Y-%m-%d').date() if day_in else d.day
        d.status = request.form.get('status', d.status)

        start_odo = request.form.get('start_odo')
        end_odo = request.form.get('end_odo')
        total_miles = request.form.get('total_miles')

        d.start_odo = int(start_odo) if start_odo else None
        d.end_odo = int(end_odo) if end_odo else None
        d.total_miles = int(total_miles) if total_miles else None

        d.start_location = request.form.get('start_location') or None
        d.trip_explanation = request.form.get('trip_explanation') or None

        segments_csv = request.form.get('segments_csv', '')
        _upsert_segments(d, segments_csv)

        if d.start_odo is not None and d.end_odo is not None and d.end_odo < d.start_odo:
            flash('End odometer cannot be less than start odometer.', 'danger')
            return redirect(url_for('work.view', day_id=d.id))

        db.session.commit()
        flash('Work Day updated.', 'success')
        return redirect(url_for('work.list'))

    segments_csv = ', '.join([s.location_name for s in d.segments])
    return render_template('work/view.html', d=d, segments_csv=segments_csv)

@work_bp.route('/end/<int:day_id>', methods=['GET', 'POST'])
@login_required
def end(day_id):
    d = WorkDay.query.get_or_404(day_id)
    if d.status != 'started':
        flash('This work day is not started.', 'danger')
        return redirect(url_for('work.list'))

    if request.method == 'POST':
        mode = request.form.get('mode', 'append')  # append | overwrite
        if mode == 'overwrite':
            _upsert_segments(d, request.form.get('segments_csv', ''))
        else:
            _append_segments(d, request.form.get('append_segments', ''))

        # End-only fields (integers or blank)
        end_odo_raw = request.form.get('end_odo')
        total_miles_raw = request.form.get('total_miles')

        d.end_odo = int(end_odo_raw) if end_odo_raw else d.end_odo
        d.total_miles = int(total_miles_raw) if total_miles_raw else d.total_miles
        d.trip_explanation = request.form.get('trip_explanation') or d.trip_explanation

        if d.start_odo is not None and d.end_odo is not None and d.end_odo < d.start_odo:
            flash('End odometer cannot be less than start odometer.', 'danger')
            return redirect(url_for('work.end', day_id=d.id))

        d.status = 'ended'
        db.session.commit()
        flash('Work day ended successfully.', 'success')
        return redirect(url_for('work.list'))

    return render_template('work/end.html', d=d)

@work_bp.route('/delete/<int:day_id>', methods=['POST'])
@login_required
def delete(day_id):
    d = WorkDay.query.get_or_404(day_id)
    db.session.delete(d)
    db.session.commit()
    flash('Work day deleted.', 'success')
    return redirect(url_for('work.list'))

@work_bp.route('/export')
@login_required
def export():
    # For simplicity, export all available months grouped by YYYY-MM
    months = (db.session.query(
        func.strftime('%Y-%m', WorkDay.day).label('ym'))
        .group_by('ym')
        .order_by('ym')
        .all())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Clear default sheet

    for (ym,) in months:
        y, m = ym.split('-')
        y, m = int(y), int(m)
        ws = wb.create_sheet(title=f'{y}-{m:02d}')
        
        headers = [
            'Date', 'Start Odo', 'End Odo', 'Total Miles',
            'Start Location', 'Segments', 'Trip Explanation',
            'Created At', 'Updated At', 'Status'
        ]
        ws.append(headers)

        start_d = date(y, m, 1)
        last_day = calendar.monthrange(y, m)[1]
        end_d = date(y, m, last_day)

        days = (WorkDay.query
            .filter(WorkDay.day >= start_d, WorkDay.day <= end_d)
            .order_by(WorkDay.day)
            .all())

        for d in days:
            ws.append([
                d.day.strftime('%Y-%m-%d'),
                d.start_odo,
                d.end_odo,
                d.compute_total_miles(),
                d.start_location or '',
                ' to '.join([s.location_name for s in d.segments]),
                d.trip_explanation or '',
                d.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                d.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                d.status
            ])

        # Auto-size columns based on content
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_len + 2

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    fname = f'work_mileage_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(bio, as_attachment=True, download_name=fname,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')